"""
Vistas para exportar reportes en PDF y Excel
"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bookings.models import Booking, Review
from services.models import Service
from accounts.models import UserProfile


def is_staff(user):
    """Verificar si el usuario es staff"""
    return user.is_staff


# ============================================================================
# PDF EXPORTS
# ============================================================================

@require_http_methods(["GET"])
@login_required
@user_passes_test(is_staff)
def export_revenue_pdf(request):
    """Exportar reporte de ingresos en PDF"""
    # Crear respuesta HTTP con tipo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte-ingresos.pdf"'
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#007bff'),
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph('Reporte de Ingresos', title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Datos generales
    now = timezone.now()
    today = now.date()
    
    # Ingresos generales
    total_revenue = Booking.objects.filter(paid=True).aggregate(Sum('total_price'))['total_price__sum'] or 0
    today_revenue = Booking.objects.filter(paid=True, payment_date__date=today).aggregate(Sum('total_price'))['total_price__sum'] or 0
    month_revenue = Booking.objects.filter(
        paid=True,
        payment_date__month=now.month,
        payment_date__year=now.year
    ).aggregate(Sum('total_price'))['total_price__sum'] or 0
    
    # Tabla de resumen
    data = [
        ['Métrica', 'Monto'],
        [f'Ingresos Totales', f'${total_revenue:.2f}'],
        [f'Ingresos Hoy', f'${today_revenue:.2f}'],
        [f'Ingresos Este Mes', f'${month_revenue:.2f}'],
    ]
    
    table = Table(data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.5*inch))
    
    # Ingresos por servicio
    story.append(Paragraph('Ingresos por Servicio', styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    service_revenue = Booking.objects.filter(paid=True).values('service__name').annotate(
        total=Sum('total_price'),
        count=Count('id')
    ).order_by('-total')
    
    service_data = [['Servicio', 'Reservas', 'Ingresos']]
    for item in service_revenue:
        service_data.append([
            item['service__name'],
            str(item['count']),
            f"${item['total']:.2f}"
        ])
    
    if len(service_data) > 1:
        service_table = Table(service_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        service_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(service_table)
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


@require_http_methods(["GET"])
@login_required
@user_passes_test(is_staff)
def export_bookings_pdf(request):
    """Exportar listado de reservas en PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte-reservas.pdf"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#007bff'),
        spaceAfter=20,
        alignment=1
    )
    story.append(Paragraph('Listado de Reservas', title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Obtener todas las reservas
    bookings = Booking.objects.select_related('user', 'service').order_by('-booking_date')[:100]
    
    # Tabla de reservas
    data = [['Usuario', 'Servicio', 'Fecha', 'Estado', 'Monto']]
    for booking in bookings:
        data.append([
            booking.user.get_full_name() or booking.user.username,
            booking.service.name,
            booking.booking_date.strftime('%d/%m/%Y'),
            booking.get_status_display(),
            f"${booking.total_price:.2f}"
        ])
    
    table = Table(data, colWidths=[1.5*inch, 1.8*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(table)
    
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


# ============================================================================
# EXCEL EXPORTS
# ============================================================================

@require_http_methods(["GET"])
@login_required
@user_passes_test(is_staff)
def export_revenue_excel(request):
    """Exportar reporte de ingresos en Excel"""
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ingresos'
    
    # Estilos
    header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = 'Reporte de Ingresos'
    ws['A1'].font = Font(bold=True, size=14, color='007bff')
    ws.merge_cells('A1:B1')
    
    # Fecha de generación
    now = timezone.now()
    today = now.date()
    ws['A2'] = f'Generado: {now.strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10)
    
    # Datos
    row = 4
    ws[f'A{row}'] = 'Métrica'
    ws[f'B{row}'] = 'Monto'
    
    for cell in [ws[f'A{row}'], ws[f'B{row}']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Ingresos generales
    total_revenue = Booking.objects.filter(paid=True).aggregate(Sum('total_price'))['total_price__sum'] or 0
    today_revenue = Booking.objects.filter(paid=True, payment_date__date=today).aggregate(Sum('total_price'))['total_price__sum'] or 0
    month_revenue = Booking.objects.filter(
        paid=True,
        payment_date__month=now.month,
        payment_date__year=now.year
    ).aggregate(Sum('total_price'))['total_price__sum'] or 0
    
    row += 1
    ws[f'A{row}'] = 'Ingresos Totales'
    ws[f'B{row}'] = total_revenue
    ws[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws[f'A{row}'] = 'Ingresos Hoy'
    ws[f'B{row}'] = today_revenue
    ws[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws[f'A{row}'] = 'Ingresos Este Mes'
    ws[f'B{row}'] = month_revenue
    ws[f'B{row}'].number_format = '$#,##0.00'
    
    # Ingresos por servicio
    row += 3
    ws[f'A{row}'] = 'Ingresos por Servicio'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    ws[f'A{row}'] = 'Servicio'
    ws[f'B{row}'] = 'Reservas'
    ws[f'C{row}'] = 'Ingresos'
    
    for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    service_revenue = Booking.objects.filter(paid=True).values('service__name').annotate(
        total=Sum('total_price'),
        count=Count('id')
    ).order_by('-total')
    
    for item in service_revenue:
        row += 1
        ws[f'A{row}'] = item['service__name']
        ws[f'B{row}'] = item['count']
        ws[f'C{row}'] = item['total']
        ws[f'C{row}'].number_format = '$#,##0.00'
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte-ingresos.xlsx"'
    wb.save(response)
    
    return response


@require_http_methods(["GET"])
@login_required
@user_passes_test(is_staff)
def export_bookings_excel(request):
    """Exportar listado de reservas en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reservas'
    
    # Estilos
    header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = 'Listado de Reservas'
    ws['A1'].font = Font(bold=True, size=14, color='007bff')
    ws.merge_cells('A1:E1')
    
    # Encabezados
    row = 3
    headers = ['Usuario', 'Servicio', 'Fecha', 'Estado', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    bookings = Booking.objects.select_related('user', 'service').order_by('-booking_date')[:200]
    
    for booking in bookings:
        row += 1
        ws.cell(row=row, column=1).value = booking.user.get_full_name() or booking.user.username
        ws.cell(row=row, column=2).value = booking.service.name
        ws.cell(row=row, column=3).value = booking.booking_date
        ws.cell(row=row, column=4).value = booking.get_status_display()
        ws.cell(row=row, column=5).value = booking.total_price
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte-reservas.xlsx"'
    wb.save(response)
    
    return response

